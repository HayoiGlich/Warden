"""Отчёт по виртуальным машинам Yandex Cloud.

Функции:
  * list_vms()          — перечень ВМ фолдера с ресурсами, снапшотами и ценой;
  * load_tariff()       — цены за сутки из шаблона (ЦПУ / ОЗУ / SSD);
  * build_report_xlsx() — заполняет шаблон выбранными ВМ и отдаёт байты xlsx.

Импорты Yandex Cloud SDK делаются лениво (внутри функций), чтобы приложение
стартовало даже без установленного пакета `yandexcloud`. Генерация xlsx требует
только openpyxl и файл-шаблон, поэтому работает независимо от доступа к облаку.
"""

from __future__ import annotations

import io
import json
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List

from backend.modules.config import settings

logger = logging.getLogger("log_analyzer")

GIB = 1024 ** 3

# platform_id -> человекочитаемое название (как в консоли YC).
PLATFORM_NAMES = {
    "standard-v1": "Intel Broadwell",
    "standard-v2": "Intel Cascade Lake",
    "standard-v3": "Intel Ice Lake",
    "standard-v3-t4": "Intel Ice Lake (T4)",
    "highfreq-v3": "Intel Ice Lake (Compute Optimized)",
    "standard-v4": "AMD Zen 4",
    "highfreq-v4": "AMD Zen 4 (Compute Optimized)",
    "gpu-standard-v1": "Intel Broadwell + GPU V100",
    "gpu-standard-v2": "Intel Cascade Lake + GPU V100",
    "gpu-standard-v3": "AMD EPYC + GPU A100",
    "gpu-standard-v3i": "Intel Ice Lake + GPU",
}

# Платформы с высокой частотой ЦПУ — тариф «Compute Optimized».
COMPUTE_OPTIMIZED_PLATFORMS = {"highfreq-v3", "highfreq-v4"}

# Явные принадлежности платформ к вендору (остальное решает эвристика по имени).
_AMD_PLATFORMS = {"standard-v4", "highfreq-v4", "gpu-standard-v3"}
_INTEL_PLATFORMS = {
    "standard-v1", "standard-v2", "standard-v3", "standard-v3-t4", "highfreq-v3",
    "gpu-standard-v1", "gpu-standard-v2", "gpu-standard-v3i",
}


def _cpu_type(platform_id: str) -> str:
    return "Compute Optimized" if platform_id in COMPUTE_OPTIMIZED_PLATFORMS else "Обычный"


def _cpu_vendor(platform_id: str) -> str:
    """Вендор ЦПУ платформы: 'intel' | 'amd'. По умолчанию Intel."""
    pid = str(platform_id or "").lower()
    if pid in _AMD_PLATFORMS:
        return "amd"
    if pid in _INTEL_PLATFORMS:
        return "intel"
    name = str(PLATFORM_NAMES.get(platform_id, platform_id) or "").lower()
    if any(m in pid or m in name for m in ("amd", "epyc", "zen")):
        return "amd"
    return "intel"


def _disk_kind(type_id: str) -> str:
    """Классификация диска по type_id: network-hdd → HDD, остальное → SSD."""
    return "hdd" if "hdd" in str(type_id or "").lower() else "ssd"


OS_TYPE_NAMES = {0: "—", 1: "Linux", 2: "Windows"}

STATUS_NAMES = {
    "PROVISIONING": "Provisioning",
    "RUNNING": "Running",
    "STOPPING": "Stopping",
    "STOPPED": "Stopped",
    "STARTING": "Starting",
    "RESTARTING": "Restarting",
    "UPDATING": "Updating",
    "ERROR": "Error",
    "CRASHED": "Crashed",
    "DELETING": "Deleting",
}


# --------------------------------------------------------------------------- #
# Пути и признак настроенности
# --------------------------------------------------------------------------- #


# Корень проекта (…/, где лежит main.py и backend/). Относительные пути к ключу
# и шаблону из .env считаем от него, а не от текущего рабочего каталога —
# иначе в Docker/сервисе (WORKDIR=/app) путь может не найтись.
_BASE_DIR = Path(__file__).resolve().parents[2]


def _resolve(value: str) -> Path:
    path = Path(value)
    return path if path.is_absolute() else (_BASE_DIR / path)


def _key_path() -> Path:
    return _resolve(settings.yc_key_file).resolve()


def _template_path() -> Path:
    return _resolve(settings.yc_template_file).resolve()


def configured() -> bool:
    """Есть ли ключ сервисного аккаунта для доступа к облаку."""
    return _key_path().is_file()


def template_ready() -> bool:
    return _template_path().is_file()


# --------------------------------------------------------------------------- #
# Тариф из шаблона
# --------------------------------------------------------------------------- #


def load_tariff() -> Dict[str, float]:
    """Цена за сутки из настроенных тарифов (вкладка «Тарифы»).

    В хранилище цены заданы за час, поэтому за сутки умножаем на 24 — как и
    формула =B2*24 в листе «Тариф» шаблона.
    """
    from backend.modules.yc_tariff import yc_tariff

    v = yc_tariff.values()

    def day(key: str) -> float:
        try:
            return round(float(v.get(key) or 0) * 24, 6)
        except (TypeError, ValueError):
            return 0.0

    def vendor(prefix: str) -> Dict[str, float]:
        return {
            "cpu_day": day(f"{prefix}_cpu_100"),   # ЦПУ 100% (обычный)
            "cpu_day_hi": day(f"{prefix}_cpu_hi"),  # ЦПУ Compute Optimized
            "ram_day": day(f"{prefix}_ram"),        # ОЗУ (обычное)
            "ram_day_hi": day(f"{prefix}_ram_hi"),  # ОЗУ (Compute Optimized)
        }

    return {
        "intel": vendor("intel"),
        "amd": vendor("amd"),
        "ssd_day": day("ssd"),  # диски общие для вендоров
        "hdd_day": day("hdd"),
    }


def _cost_day(
    tariff: Dict[str, Any], vendor: str, cpu_type: str, cores, ram_gb, ssd_gb, hdd_gb
) -> float:
    """Стоимость ВМ в сутки: ЦПУ+ОЗУ по вендору и типу платформы, диски раздельно."""
    vt = tariff.get(vendor) or tariff.get("intel") or {}
    is_hi = cpu_type == "Compute Optimized"
    cpu_price = vt.get("cpu_day_hi", 0) if is_hi else vt.get("cpu_day", 0)
    ram_price = vt.get("ram_day_hi", 0) if is_hi else vt.get("ram_day", 0)
    return _round2(
        (cores or 0) * cpu_price
        + (ram_gb or 0) * ram_price
        + (ssd_gb or 0) * tariff.get("ssd_day", 0)
        + (hdd_gb or 0) * tariff.get("hdd_day", 0)
    )


# --------------------------------------------------------------------------- #
# Актуальные цены из Yandex Cloud Billing API (кнопка «Вставить цены»)
# --------------------------------------------------------------------------- #

# Поле тарифа -> SKU Compute Cloud (сервис dn22pas77ftg9h3f2djj), платформа
# Intel Ice Lake — та, на которой работают ВМ фолдера. Цены в SKU заданы за час
# (core*hour / gbyte*hour), что совпадает с единицей полей тарифа.
YC_PRICE_SKUS = {
    # Intel Ice Lake
    "intel_cpu_100": "dn2k3vqlk9snp1jv351u",  # Regular Intel Ice Lake, 100% vCPU
    "intel_cpu_50": "dn2f0q0d6gtpcom4b1p6",   # Regular Intel Ice Lake, 50% vCPU
    "intel_cpu_hi": "dn2lag3718gm9oq8dus2",   # Intel Ice Lake (Compute Optimized), 100%
    "intel_ram": "dn2ilq72mjc3bej6j74p",      # Regular Intel Ice Lake, RAM
    "intel_ram_hi": "dn23hq90a5khr3o6fivm",   # Intel Ice Lake (Compute Optimized), RAM
    # AMD Zen 4
    "amd_cpu_100": "dn28kn5h601tc7lk5fbu",    # Regular AMD Zen 4, 100% vCPU
    "amd_cpu_50": "dn292qoem8pkl2rig4i7",     # Regular AMD Zen 4, 50% vCPU
    "amd_cpu_hi": "dn2bnom85ie58bpvmtmn",     # AMD Zen 4 (Compute Optimized), 100% vCPU
    "amd_ram": "dn29sa4d441spg8aokdn",        # Regular AMD Zen 4, RAM
    "amd_ram_hi": "dn2dq9mqiklrm87pc1h2",     # AMD Zen 4 (Compute Optimized), RAM
    # Диски (общие)
    "ssd": "dn27ajm6m8mnfcshbi61",            # Fast network drive (SSD) = network-ssd
    "ssd_io": "dn2bl3v71k1mej7andmc",         # Ultra high-speed (SSD, 3 replicas)
    "hdd": "dn2al287u6jr3a710u8g",            # Standard disk drive (HDD) = network-hdd
}

_IAM_TOKEN_URL = "https://iam.api.cloud.yandex.net/iam/v1/tokens"
_BILLING_SKU_URL = "https://billing.api.cloud.yandex.net/billing/v1/skus/{sku}"


def _iam_token() -> str:
    """IAM-токен по ключу сервисного аккаунта (JWT -> IAM API)."""
    import time

    import jwt
    import requests

    key = _load_key_data()
    now = int(time.time())
    encoded = jwt.encode(
        {
            "aud": _IAM_TOKEN_URL,
            "iss": key["service_account_id"],
            "iat": now,
            "exp": now + 3600,
        },
        key["private_key"],
        algorithm="PS256",
        headers={"kid": key["id"]},
    )
    resp = requests.post(_IAM_TOKEN_URL, json={"jwt": encoded}, timeout=20)
    resp.raise_for_status()
    return resp.json()["iamToken"]


def _latest_street_price(sku: dict) -> float:
    versions = [
        v for v in sku.get("pricingVersions", []) if v.get("type") == "STREET_PRICE"
    ]
    if not versions:
        return 0.0
    rates = versions[-1].get("pricingExpressions", [{}])[0].get("rates", [{}])
    try:
        return round(float(rates[0].get("unitPrice") or 0), 6)
    except (TypeError, ValueError):
        return 0.0


def fetch_tariff_prices() -> Dict[str, Any]:
    """Актуальные цены за час (₽) из Billing API по фиксированным SKU.

    Возвращает {"prices": {поле: цена}, "as_of": ISO-дата действия}.
    """
    import requests

    token = _iam_token()
    headers = {"Authorization": f"Bearer {token}"}
    prices: Dict[str, float] = {}
    as_of = ""
    for field, sku_id in YC_PRICE_SKUS.items():
        resp = requests.get(
            _BILLING_SKU_URL.format(sku=sku_id),
            params={"currency": "RUB"},
            headers=headers,
            timeout=25,
        )
        resp.raise_for_status()
        sku = resp.json()
        prices[field] = _latest_street_price(sku)
        versions = [
            v for v in sku.get("pricingVersions", []) if v.get("type") == "STREET_PRICE"
        ]
        if versions and versions[-1].get("effectiveTime"):
            eff = versions[-1]["effectiveTime"]
            if eff > as_of:
                as_of = eff
    return {"prices": prices, "as_of": as_of}


# --------------------------------------------------------------------------- #
# Листинг ВМ
# --------------------------------------------------------------------------- #


def _load_key_data() -> dict:
    key_path = _key_path()
    if not key_path.is_file():
        raise RuntimeError(f"Не найден ключ сервисного аккаунта: {key_path}")
    with key_path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _init_sdk():
    try:
        import yandexcloud
    except ImportError as exc:  # pragma: no cover - зависит от окружения
        raise RuntimeError(
            "Библиотека yandexcloud не установлена. "
            "Установите зависимости (requirements-web.txt)."
        ) from exc

    return yandexcloud.SDK(service_account_key=_load_key_data())


def _round2(value: float) -> float:
    return round(float(value or 0), 2)


def _fmt_created(ts) -> str:
    """protobuf Timestamp -> ДД.ММ.ГГГГ. Пусто, если даты нет."""
    seconds = int(getattr(ts, "seconds", 0) or 0)
    if not seconds:
        return ""
    return datetime.fromtimestamp(seconds, tz=timezone.utc).strftime("%d.%m.%Y")


def list_vms() -> List[Dict[str, Any]]:
    """Перечень ВМ фолдера: ресурсы, размеры дисков/снапшотов и расчётная цена."""
    sdk = _init_sdk()

    from yandex.cloud.compute.v1.instance_pb2 import Instance
    from yandex.cloud.compute.v1.instance_service_pb2 import ListInstancesRequest
    from yandex.cloud.compute.v1.instance_service_pb2_grpc import InstanceServiceStub
    from yandex.cloud.compute.v1.disk_service_pb2 import GetDiskRequest
    from yandex.cloud.compute.v1.disk_service_pb2_grpc import DiskServiceStub
    from yandex.cloud.compute.v1.image_service_pb2 import GetImageRequest
    from yandex.cloud.compute.v1.image_service_pb2_grpc import ImageServiceStub
    from yandex.cloud.compute.v1.snapshot_service_pb2 import ListSnapshotsRequest
    from yandex.cloud.compute.v1.snapshot_service_pb2_grpc import SnapshotServiceStub

    inst_stub = sdk.client(InstanceServiceStub)
    disk_stub = sdk.client(DiskServiceStub)
    image_stub = sdk.client(ImageServiceStub)
    snap_stub = sdk.client(SnapshotServiceStub)

    fid = settings.yc_folder_id
    zone_filter = (settings.yc_report_zone or "").strip()
    tariff = load_tariff()

    # disk_id -> (size_bytes, source_image_id)
    disk_cache: Dict[str, tuple] = {}
    # image_id -> "Linux" | "Windows" | "—"
    image_os_cache: Dict[str, str] = {}
    # disk_id -> суммарный размер снапшотов в байтах
    snap_cache: Dict[str, int] = {}

    def disk_meta(disk_id: str) -> tuple:
        if not disk_id:
            return (0, "", "")
        if disk_id not in disk_cache:
            try:
                d = disk_stub.Get(GetDiskRequest(disk_id=disk_id))
                disk_cache[disk_id] = (
                    int(d.size),
                    getattr(d, "source_image_id", "") or "",
                    getattr(d, "type_id", "") or "",
                )
            except Exception as exc:
                logger.warning("YC: не удалось получить диск %s: %s", disk_id, exc)
                disk_cache[disk_id] = (0, "", "")
        return disk_cache[disk_id]

    def snapshots_bytes(disk_id: str) -> int:
        if not disk_id:
            return 0
        if disk_id not in snap_cache:
            total = 0
            try:
                resp = snap_stub.List(
                    ListSnapshotsRequest(
                        folder_id=fid,
                        filter=f'source_disk_id = "{disk_id}"',
                    )
                )
                for snap in resp.snapshots:
                    total += int(
                        getattr(snap, "storage_size", 0)
                        or getattr(snap, "disk_size", 0)
                        or 0
                    )
            except Exception as exc:
                logger.warning("YC: снапшоты диска %s недоступны: %s", disk_id, exc)
            snap_cache[disk_id] = total
        return snap_cache[disk_id]

    def resolve_os(image_id: str) -> str:
        if not image_id:
            return "—"
        if image_id in image_os_cache:
            return image_os_cache[image_id]
        label = "—"
        try:
            img = image_stub.Get(GetImageRequest(image_id=image_id))
            os_type = int(getattr(getattr(img, "os", None), "type", 0) or 0)
            label = OS_TYPE_NAMES.get(os_type, "—")
            if label == "—":
                label = _detect_os_from_family(getattr(img, "family", ""))
        except Exception as exc:
            logger.warning("YC: образ %s недоступен: %s", image_id, exc)
        image_os_cache[image_id] = label
        return label

    rows: List[Dict[str, Any]] = []
    page_token = ""
    while True:
        resp = inst_stub.List(
            ListInstancesRequest(folder_id=fid, page_size=100, page_token=page_token)
        )
        for inst in resp.instances:
            if zone_filter and inst.zone_id != zone_filter:
                continue

            disk_ids: List[str] = []
            boot_disk_id = ""
            if getattr(inst, "boot_disk", None) and getattr(inst.boot_disk, "disk_id", ""):
                boot_disk_id = inst.boot_disk.disk_id
                disk_ids.append(boot_disk_id)
            for sd in getattr(inst, "secondary_disks", []):
                if getattr(sd, "disk_id", ""):
                    disk_ids.append(sd.disk_id)

            ssd_bytes = 0
            hdd_bytes = 0
            total_snap_bytes = 0
            for disk_id in disk_ids:
                size_b, _img, type_id = disk_meta(disk_id)
                if _disk_kind(type_id) == "hdd":
                    hdd_bytes += size_b
                else:
                    ssd_bytes += size_b
                total_snap_bytes += snapshots_bytes(disk_id)

            boot_image_id = disk_meta(boot_disk_id)[1] if boot_disk_id else ""

            cores = int(inst.resources.cores)
            core_fraction = int(inst.resources.core_fraction)
            ram_gb = _round2(int(inst.resources.memory) / GIB)
            ssd_gb = _round2(ssd_bytes / GIB)
            hdd_gb = _round2(hdd_bytes / GIB)
            snap_gb = _round2(total_snap_bytes / GIB)

            cpu_type = _cpu_type(inst.platform_id)
            cpu_vendor = _cpu_vendor(inst.platform_id)
            cost_day = _cost_day(
                tariff, cpu_vendor, cpu_type, cores, ram_gb, ssd_gb, hdd_gb
            )
            cost_year = _round2(cost_day * 365)

            rows.append(
                {
                    "id": inst.id,
                    "name": inst.name,
                    "created_at": _fmt_created(getattr(inst, "created_at", None)),
                    "zone_id": inst.zone_id,
                    "status": _fmt_status(Instance, inst.status),
                    "platform": PLATFORM_NAMES.get(inst.platform_id, inst.platform_id or ""),
                    "platform_id": inst.platform_id,
                    "cpu_type": cpu_type,
                    "cpu_vendor": cpu_vendor,
                    "os": resolve_os(boot_image_id),
                    "cores": cores,
                    "core_fraction": core_fraction,
                    "ram_gb": ram_gb,
                    "ssd_gb": ssd_gb,
                    "hdd_gb": hdd_gb,
                    "disk_gb": _round2((ssd_bytes + hdd_bytes) / GIB),
                    "snapshots_gb": snap_gb,
                    "cost_day": cost_day,
                    "cost_year": cost_year,
                }
            )

        page_token = resp.next_page_token
        if not page_token:
            break

    rows.sort(key=lambda r: str(r["name"]).lower())
    return rows


# --------------------------------------------------------------------------- #
# Кэш листинга ВМ (обращение к облаку медленное — не дёргаем на каждый заход)
# --------------------------------------------------------------------------- #

# {"data": [...], "ts": epoch_seconds}. Живёт в памяти процесса до обновления.
_vms_cache: Dict[str, Any] = {"data": None, "ts": 0.0}


def get_vms(force: bool = False) -> tuple:
    """Отдаёт (список_ВМ, время_загрузки_epoch). Из кэша, если он есть и не
    запрошено принудительное обновление."""
    if not force and _vms_cache["data"] is not None:
        return _vms_cache["data"], _vms_cache["ts"]
    import time

    data = list_vms()
    _vms_cache["data"] = data
    _vms_cache["ts"] = time.time()
    return data, _vms_cache["ts"]


def cached_at() -> float:
    return float(_vms_cache["ts"] or 0.0)


def has_cache() -> bool:
    return _vms_cache["data"] is not None


def recompute_cached_costs() -> None:
    """Пересчитывает стоимость в кэше по актуальному тарифу — без похода в облако
    (нужно после изменения тарифов, чтобы не перезагружать все ВМ)."""
    if _vms_cache["data"] is None:
        return
    tariff = load_tariff()
    for vm in _vms_cache["data"]:
        vm["cost_day"] = _cost_day(
            tariff,
            vm.get("cpu_vendor", "intel"),
            vm.get("cpu_type", ""),
            vm.get("cores", 0),
            vm.get("ram_gb", 0),
            vm.get("ssd_gb", 0),
            vm.get("hdd_gb", 0),
        )
        vm["cost_year"] = _round2(vm["cost_day"] * 365)


def _fmt_status(instance_cls, status_int: int) -> str:
    try:
        name = instance_cls.Status.Name(status_int)
    except Exception:
        return str(status_int)
    return STATUS_NAMES.get(name, name.capitalize())


def _detect_os_from_family(family: str) -> str:
    f = (family or "").lower()
    if not f:
        return "—"
    if "win" in f:
        return "Windows"
    linux_markers = (
        "ubuntu", "debian", "centos", "rocky", "almalinux", "fedora", "rhel",
        "redhat", "suse", "opensuse", "linux", "astra", "alt", "oracle", "amazon",
    )
    if any(m in f for m in linux_markers):
        return "Linux"
    return "—"


# --------------------------------------------------------------------------- #
# Генерация отчёта из шаблона
# --------------------------------------------------------------------------- #


def build_report_xlsx(rows: List[Dict[str, Any]]) -> bytes:
    """Заполняет шаблон выбранными ВМ и возвращает содержимое xlsx (байты).

    Листы «Список ВМ» и «Снимки» — исходные данные, «Общий» — расчёт стоимости.
    Цена считается по вендору (Intel/AMD) и типу платформы (обычный / Compute
    Optimized), диски (SSD/HDD) раздельно. Лист «Тариф» переписывается в
    справочную таблицу цен Intel/AMD, по которой шёл расчёт.
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    from backend.modules.yc_tariff import yc_tariff

    path = _template_path()
    if not path.is_file():
        raise RuntimeError(f"Не найден шаблон отчёта: {path}")

    wb = load_workbook(path, data_only=False)
    ws_list = wb["Список ВМ"]
    ws_snap = wb["Снимки"]
    ws_total = wb["Общий"]
    ws_tariff = wb["Тариф"]

    tariff = load_tariff()          # за сутки, по вендорам — для расчёта
    tv = yc_tariff.values()         # за час — для справочного листа «Тариф»

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E2EFDA")

    list_headers = [
        "Имя ВМ", "Дата создания", "Платформа", "Тип ЦПУ", "ЦПУ, шт.", "ОЗУ, Гб",
        "SSD, Гб", "HDD, Гб",
    ]
    snap_headers = ["Имя ВМ", "Снимки, Гб"]
    total_headers = [
        "Имя ВМ", "Дата создания", "Платформа", "Тип ЦПУ", "ЦПУ, шт.", "ОЗУ, Гб",
        "SSD, Гб", "HDD, Гб", "Снимки, Гб", "Итого за ВМ в день, ₽",
        "Итого за ВМ в год, ₽", "Итого за всё в год, ₽",
    ]

    def reset(ws, headers):
        if ws.max_row >= 1:
            ws.delete_rows(1, ws.max_row)
        for col, title in enumerate(headers, start=1):
            cell = ws.cell(1, col, title)
            cell.font = header_font
            cell.fill = header_fill

    reset(ws_list, list_headers)
    reset(ws_snap, snap_headers)
    reset(ws_total, total_headers)

    for idx, row in enumerate(rows, start=2):
        name = str(row.get("name") or "")
        created_at = str(row.get("created_at") or "")
        platform = str(row.get("platform") or "")
        cpu_type = str(row.get("cpu_type") or "Обычный")
        vendor = str(row.get("cpu_vendor") or "intel")
        cores = row.get("cores") or 0
        ram_gb = row.get("ram_gb") or 0
        ssd_gb = row.get("ssd_gb") or 0
        hdd_gb = row.get("hdd_gb") or 0
        snap_gb = row.get("snapshots_gb") or 0

        ws_list.cell(idx, 1, name)
        ws_list.cell(idx, 2, created_at)
        ws_list.cell(idx, 3, platform)
        ws_list.cell(idx, 4, cpu_type)
        ws_list.cell(idx, 5, cores)
        ws_list.cell(idx, 6, ram_gb)
        ws_list.cell(idx, 7, ssd_gb)
        ws_list.cell(idx, 8, hdd_gb)

        ws_snap.cell(idx, 1, name)
        ws_snap.cell(idx, 2, snap_gb)

        day = _cost_day(tariff, vendor, cpu_type, cores, ram_gb, ssd_gb, hdd_gb)
        year = _round2(day * 365)

        # A..H — зеркало «Список ВМ», I — снимки, J/K — рассчитанная стоимость
        # (числами: тариф зависит от вендора построчно, поэтому расчёт делаем
        # здесь, а не единой формулой на весь лист).
        for col, letter in enumerate("ABCDEFGH", start=1):
            ws_total.cell(idx, col, f"='Список ВМ'!{letter}{idx}")
        ws_total.cell(idx, 9, f"='Снимки'!B{idx}")
        ws_total.cell(idx, 10, day)
        ws_total.cell(idx, 11, year)

    if rows:
        last = len(rows) + 1
        ws_total.cell(2, 12, f"=SUM(K2:K{last})")

    # Лист «Тариф» — справочная таблица цен за час, по которой шёл расчёт.
    if ws_tariff.max_row >= 1:
        ws_tariff.delete_rows(1, ws_tariff.max_row)
    tariff_table = [
        ("Тариф (цена за час, ₽)", "Intel", "AMD"),
        ("ЦПУ 100%", tv.get("intel_cpu_100"), tv.get("amd_cpu_100")),
        ("ЦПУ 50%", tv.get("intel_cpu_50"), tv.get("amd_cpu_50")),
        ("ЦПУ Compute Optimized", tv.get("intel_cpu_hi"), tv.get("amd_cpu_hi")),
        ("ОЗУ (за ГБ)", tv.get("intel_ram"), tv.get("amd_ram")),
        ("ОЗУ Compute Optimized (за ГБ)", tv.get("intel_ram_hi"), tv.get("amd_ram_hi")),
        ("SSD (за ГБ)", tv.get("ssd"), tv.get("ssd")),
        ("SSD IO (за ГБ)", tv.get("ssd_io"), tv.get("ssd_io")),
        ("HDD (за ГБ)", tv.get("hdd"), tv.get("hdd")),
    ]
    for r, (a, b, c) in enumerate(tariff_table, start=1):
        ca = ws_tariff.cell(r, 1, a)
        ws_tariff.cell(r, 2, b)
        ws_tariff.cell(r, 3, c)
        if r == 1:
            for col in (1, 2, 3):
                ws_tariff.cell(r, col).font = header_font
                ws_tariff.cell(r, col).fill = header_fill
    ws_tariff.column_dimensions["A"].width = 32
    ws_tariff.column_dimensions["B"].width = 12
    ws_tariff.column_dimensions["C"].width = 12

    # Ширина колонок для читабельности + закреплённая шапка.
    for ws, headers in (
        (ws_list, list_headers),
        (ws_snap, snap_headers),
        (ws_total, total_headers),
    ):
        for col_idx, title in enumerate(headers, start=1):
            letter = ws.cell(1, col_idx).column_letter
            ws.column_dimensions[letter].width = max(12, min(32, len(title) + 3))
        ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
