"""Тарифы Yandex Cloud (цены за час, ₽) для расчёта стоимости ВМ.

Цены ЦПУ и ОЗУ раздельны для Intel и AMD (у платформ разный прайс); диски
общие. Хранится в app_setting под ключом `yc_tariff` (JSON). По умолчанию Intel
берётся из листа «Тариф» шаблона, AMD — зеркалит Intel, пока не заданы явно.
Редактирует админ на панели «Тарифы» в отчёте по ВМ.
"""

from __future__ import annotations

import json
import logging

logger = logging.getLogger("log_analyzer")

SETTING_KEY = "yc_tariff"

# Поля ЦПУ/ОЗУ по вендорам + общие диски.
CPU_RAM_FIELDS = ("cpu_100", "cpu_50", "cpu_hi", "ram", "ram_hi")
DISK_FIELDS = ("ssd", "ssd_io", "hdd")

FIELDS = tuple(
    [f"intel_{f}" for f in CPU_RAM_FIELDS]
    + [f"amd_{f}" for f in CPU_RAM_FIELDS]
    + list(DISK_FIELDS)
)

# Ячейки листа «Тариф» шаблона (строка 2 — цена за час) для дефолтов Intel + диски.
_TEMPLATE_CELLS = {
    "intel_cpu_100": "B2",
    "intel_cpu_50": "C2",
    "intel_cpu_hi": "D2",
    "intel_ram": "E2",
    "intel_ram_hi": "F2",
    "ssd": "G2",
    "ssd_io": "H2",
    "hdd": "I2",
}

# Старые (одновендорные) ключи -> новые Intel-ключи, для совместимости с
# сохранёнными ранее тарифами.
_LEGACY_MAP = {
    "cpu_100": "intel_cpu_100",
    "cpu_50": "intel_cpu_50",
    "cpu_hi": "intel_cpu_hi",
    "ram": "intel_ram",
    "ram_hi": "intel_ram_hi",
}


def _num(value) -> float:
    try:
        return round(float(value), 6)
    except (TypeError, ValueError):
        return 0.0


def _clean(data) -> dict:
    """Приводит вход к полному набору полей.

    Понимает старый формат (cpu_100 -> intel_cpu_100). AMD, если не задан,
    зеркалит Intel — чтобы расчёт не занулялся до явной настройки/«Вставить цены».
    """
    data = data or {}
    out: dict[str, float] = {}
    for field in FIELDS:
        if field in data:
            out[field] = _num(data.get(field))
        elif field in _LEGACY_MAP.values():
            # новый Intel-ключ отсутствует — пробуем старый одновендорный
            legacy = next((k for k, v in _LEGACY_MAP.items() if v == field), None)
            out[field] = _num(data.get(legacy)) if legacy in (data or {}) else 0.0
        else:
            out[field] = 0.0
    # AMD зеркалит Intel там, где AMD не задан (0), а Intel есть.
    for f in CPU_RAM_FIELDS:
        if not out.get(f"amd_{f}") and out.get(f"intel_{f}"):
            out[f"amd_{f}"] = out[f"intel_{f}"]
    return out


class YcTariff:
    def __init__(self) -> None:
        self._values: dict | None = None

    def _template_defaults(self) -> dict:
        """Цены по умолчанию — Intel из листа «Тариф» шаблона, AMD зеркалит Intel."""
        raw: dict[str, float] = {}
        try:
            from openpyxl import load_workbook

            from backend.modules.yc_report import _template_path

            path = _template_path()
            if path.is_file():
                ws = load_workbook(path, data_only=False)["Тариф"]
                raw = {key: _num(ws[cell].value) for key, cell in _TEMPLATE_CELLS.items()}
        except Exception:
            logger.exception("YC tariff: не удалось прочитать шаблон")
        return _clean(raw)

    async def load(self) -> None:
        try:
            from backend.modules.app_db import app_db

            if not app_db.ready:
                return
            raw = (await app_db.get_setting(SETTING_KEY, "")).strip()
            if raw:
                self._values = _clean(json.loads(raw))
                logger.info("YC tariff: загружено из БД")
            else:
                self._values = self._template_defaults()
                logger.info("YC tariff: используются значения из шаблона")
        except Exception:
            logger.exception("YC tariff: ошибка загрузки")

    def values(self) -> dict:
        if self._values is None:
            return self._template_defaults()
        return dict(self._values)

    async def save(self, data: dict) -> dict:
        self._values = _clean(data)
        from backend.modules.app_db import app_db

        await app_db.set_setting(
            SETTING_KEY, json.dumps(self._values, ensure_ascii=False)
        )
        logger.info("YC tariff: сохранено")
        return dict(self._values)


yc_tariff = YcTariff()
