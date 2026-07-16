from __future__ import annotations

import json
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, List, Set

import yandexcloud
from yandex.cloud.compute.v1.instance_pb2 import Instance
from yandex.cloud.compute.v1.instance_service_pb2 import ListInstancesRequest
from yandex.cloud.compute.v1.instance_service_pb2_grpc import InstanceServiceStub
from yandex.cloud.compute.v1.disk_service_pb2 import GetDiskRequest
from yandex.cloud.compute.v1.disk_service_pb2_grpc import DiskServiceStub
from yandex.cloud.compute.v1.image_service_pb2 import GetImageRequest
from yandex.cloud.compute.v1.image_service_pb2_grpc import ImageServiceStub


# ==================== КОНФИГУРАЦИЯ ====================
FOLDER_ID = "b1g63k911vl1iaaku9eh"
KEY_FILE = "authorized_key.json"
TARGET_ZONE_FOR_REPORT = ""  # если нужно только 1 зона; поставь "" чтобы не фильтровать

OUTPUT_XLSX = "yc_vm_report.xlsx"
# ====================


# Маппинг platform_id -> человекочитаемое название (как в консоли YC).
PLATFORM_NAMES = {
    "standard-v1": "Intel Broadwell",
    "standard-v2": "Intel Cascade Lake",
    "standard-v3": "Intel Ice Lake",
    "standard-v3-t4": "Intel Ice Lake (T4)",
    "highfreq-v3": "Intel Ice Lake (Compute Optimized)",
    "gpu-standard-v1": "Intel Broadwell + GPU V100",
    "gpu-standard-v2": "Intel Cascade Lake + GPU V100",
    "gpu-standard-v3": "AMD EPYC + GPU A100",
    "gpu-standard-v3i": "Intel Ice Lake + GPU",
}


STATUS_NAMES = {
    "PROVISIONING": "Provisioning",
    "RUNNING": "Running",
    "STOPPING": "Stopping",
    "STOPPED": "Stopped",
    "STARTING": "Starting",
    "RESTARTING": "Restarting",
    "UPDATING": "Updating",
    "ERROR": "Error",
    "CRASHED": "Crashed",
    "DELETING": "Deleting",
}


def load_key_data() -> dict:
    with open(KEY_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def init_sdk(key_data: dict):
    """Инициализирует SDK Yandex Cloud."""
    return yandexcloud.SDK(service_account_key=key_data)


def get_iam_token(key_data: dict) -> str:
    """Получает IAM-токен по ключу сервисного аккаунта (JWT -> IAM API)."""
    import jwt  # из pyjwt
    import requests

    now = int(time.time())
    payload = {
        "aud": "https://iam.api.cloud.yandex.net/iam/v1/tokens",
        "iss": key_data["service_account_id"],
        "iat": now,
        "exp": now + 3600,
    }
    encoded = jwt.encode(
        payload,
        key_data["private_key"],
        algorithm="PS256",
        headers={"kid": key_data["id"]},
    )
    r = requests.post(
        "https://iam.api.cloud.yandex.net/iam/v1/tokens",
        json={"jwt": encoded},
        timeout=15,
    )
    r.raise_for_status()
    return r.json()["iamToken"]


def fmt_date(ts) -> str:
    if not ts or not getattr(ts, "seconds", 0):
        return ""
    dt = datetime.fromtimestamp(ts.seconds, tz=timezone.utc)
    return dt.strftime("%d.%m.%Y")


def fmt_platform(platform_id: str) -> str:
    return PLATFORM_NAMES.get(platform_id, platform_id or "")


def fmt_status(status_int: int) -> str:
    try:
        name = Instance.Status.Name(status_int)
    except Exception:
        return str(status_int)
    return STATUS_NAMES.get(name, name.capitalize())


def fmt_fraction(core_fraction: int):
    """core_fraction в YC хранится в процентах (100 = полное ядро). Возвращаем число."""
    frac = (core_fraction or 0) / 100.0
    if frac == int(frac):
        return int(frac)
    return round(frac, 2)


def fmt_ram(memory_bytes: int) -> str:
    gib = memory_bytes / (1024 ** 3)
    if gib == int(gib):
        return f"{int(gib)} ГБ"
    return f"{round(gib, 2)} ГБ"


def fmt_disk_size(total_bytes: int) -> str:
    if total_bytes <= 0:
        return "0 ГБ"
    gib = total_bytes / (1024 ** 3)
    if gib >= 1024:
        tib = gib / 1024
        s = f"{tib:.2f}".rstrip("0").rstrip(".")
        return f"{s} ТБ"
    if gib == int(gib):
        return f"{int(gib)} ГБ"
    s = f"{gib:.2f}".rstrip("0").rstrip(".")
    return f"{s} ГБ"


OS_TYPE_NAMES = {
    0: "—",        # OS_TYPE_UNSPECIFIED
    1: "Linux",
    2: "Windows",
}


def detect_os_from_family(family: str) -> str:
    """Резервная эвристика: определяем ОС по имени family образа."""
    f = (family or "").lower()
    if not f:
        return "—"
    if "win" in f:
        return "Windows"
    linux_markers = (
        "ubuntu", "debian", "centos", "rocky", "almalinux", "fedora",
        "rhel", "redhat", "suse", "opensuse", "linux", "astra",
        "alt", "oracle", "amazon",
    )
    if any(m in f for m in linux_markers):
        return "Linux"
    return "—"


@dataclass
class VmRow:
    created_at: str
    name: str
    cloud_backup: str
    status: str
    platform: str
    os: str
    vcpu: int
    vcpu_fraction: Any
    ram: str
    disk_size: str


def list_all_instances(instance_stub: InstanceServiceStub, folder_id: str) -> Iterable:
    """Пагинированный листинг ВМ по folder_id."""
    page_token = ""
    while True:
        resp = instance_stub.List(
            ListInstancesRequest(
                folder_id=folder_id, page_size=100, page_token=page_token
            )
        )
        for inst in resp.instances:
            yield inst
        page_token = resp.next_page_token
        if not page_token:
            break


def get_backed_up_instance_ids(key_data: dict, folder_id: str) -> Set[str]:
    """Возвращает множество ID ВМ, у которых подключён Cloud Backup.

    Идём через REST API: gRPC-стаб сервиса Cloud Backup не зарегистрирован
    в карте сервисов yandexcloud SDK и через sdk.client(...) не работает.
    """
    try:
        import requests
    except ImportError:
        print("Библиотека requests не установлена — Cloud Backup пропущен.")
        return set()

    try:
        token = get_iam_token(key_data)
    except Exception as e:
        print(f"Внимание: не удалось получить IAM-токен: {e}")
        return set()

    ids: Set[str] = set()
    url = "https://backup.api.cloud.yandex.net/backup/v1/resources"
    headers = {"Authorization": f"Bearer {token}"}
    params: Dict[str, Any] = {"folderId": folder_id, "pageSize": 1000}

    try:
        while True:
            r = requests.get(url, params=params, headers=headers, timeout=30)
            r.raise_for_status()
            data = r.json()
            for res in data.get("resources", []):
                cid = (
                    res.get("computeInstanceId")
                    or res.get("instanceId")
                    or res.get("id")
                )
                if cid:
                    ids.add(cid)
            next_token = data.get("nextPageToken")
            if not next_token:
                break
            params["pageToken"] = next_token
    except Exception as e:
        print(f"Внимание: не удалось получить список бэкапов: {e}")
    return ids


def build_report_rows(sdk, key_data: dict) -> List[VmRow]:
    instance_stub = sdk.client(InstanceServiceStub)
    disk_stub = sdk.client(DiskServiceStub)
    image_stub = sdk.client(ImageServiceStub)

    backed_up_ids = get_backed_up_instance_ids(key_data, FOLDER_ID)

    # disk_id -> (size_bytes, source_image_id)
    disk_cache: Dict[str, tuple] = {}
    # image_id -> "Linux" | "Windows" | "—"
    image_os_cache: Dict[str, str] = {}

    def resolve_os_for_image(image_id: str) -> str:
        if not image_id:
            return "—"
        if image_id in image_os_cache:
            return image_os_cache[image_id]
        os_label = "—"
        try:
            img = image_stub.Get(GetImageRequest(image_id=image_id))
            os_type = int(getattr(getattr(img, "os", None), "type", 0) or 0)
            os_label = OS_TYPE_NAMES.get(os_type, "—")
            if os_label == "—":
                os_label = detect_os_from_family(getattr(img, "family", ""))
        except Exception as e:
            print(f"Внимание: не удалось получить образ {image_id}: {e}")
        image_os_cache[image_id] = os_label
        return os_label

    rows: List[VmRow] = []

    for inst in list_all_instances(instance_stub, FOLDER_ID):
        if TARGET_ZONE_FOR_REPORT and inst.zone_id != TARGET_ZONE_FOR_REPORT:
            continue

        cpu = int(inst.resources.cores)
        core_fraction = int(inst.resources.core_fraction)
        memory_bytes = int(inst.resources.memory)

        total_disk_bytes = 0
        disk_ids: list[str] = []
        boot_disk_id = ""

        if getattr(inst, "boot_disk", None) and getattr(inst.boot_disk, "disk_id", ""):
            boot_disk_id = inst.boot_disk.disk_id
            disk_ids.append(boot_disk_id)
        for sd in getattr(inst, "secondary_disks", []):
            if getattr(sd, "disk_id", ""):
                disk_ids.append(sd.disk_id)

        for disk_id in disk_ids:
            if disk_id not in disk_cache:
                try:
                    d = disk_stub.Get(GetDiskRequest(disk_id=disk_id))
                    disk_cache[disk_id] = (
                        int(d.size),
                        getattr(d, "source_image_id", "") or "",
                    )
                except Exception as e:
                    print(f"Внимание: не удалось получить диск {disk_id}: {e}")
                    disk_cache[disk_id] = (0, "")
            total_disk_bytes += disk_cache[disk_id][0]

        boot_image_id = disk_cache.get(boot_disk_id, (0, ""))[1] if boot_disk_id else ""
        os_label = resolve_os_for_image(boot_image_id)

        rows.append(
            VmRow(
                created_at=fmt_date(inst.created_at),
                name=inst.name,
                cloud_backup="Подключён" if inst.id in backed_up_ids else "Не подключён",
                status=fmt_status(inst.status),
                platform=fmt_platform(inst.platform_id),
                os=os_label,
                vcpu=cpu,
                vcpu_fraction=fmt_fraction(core_fraction),
                ram=fmt_ram(memory_bytes),
                disk_size=fmt_disk_size(total_disk_bytes),
            )
        )

    rows.sort(key=lambda r: r.name.lower())
    return rows


def write_xlsx(rows: List[VmRow], path: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "VMs"

    headers = [
        "Дата создания",
        "Имя",
        "Cloud Backup",
        "Статус",
        "Платформа",
        "ОС",
        "vCPU",
        "Доля vCPU",
        "RAM",
        "Размер диска",
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E2EFDA")  # светло-зелёный как на скрине
    header_align = Alignment(horizontal="left", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for r in rows:
        ws.append(
            [
                r.created_at,
                r.name,
                r.cloud_backup,
                r.status,
                r.platform,
                r.os,
                r.vcpu,
                r.vcpu_fraction,
                r.ram,
                r.disk_size,
            ]
        )

    # Подгоняем ширину колонок под содержимое
    for col_idx in range(1, len(headers) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = max(
            (len(str(c.value)) for c in ws[col_letter] if c.value is not None),
            default=10,
        )
        ws.column_dimensions[col_letter].width = max(10, min(40, max_len + 2))

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(path)


def main() -> None:
    key_data = load_key_data()
    sdk = init_sdk(key_data)
    rows = build_report_rows(sdk, key_data)
    write_xlsx(rows, OUTPUT_XLSX)
    print(f"OK: {len(rows)} VM rows written to {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
